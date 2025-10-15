#!/usr/bin/env python3
"""
Generate sample Excel workbook files for testing Roman's Rater application.
"""

from openpyxl import Workbook
from pathlib import Path


def create_cover_whale_rater_file():
    """Create the 2025 Cover Whale Rater AL only version.xlsx file."""
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Sheet 1: Rating Plan by State
    ws_rating = wb.create_sheet("Rating Plan by State")
    ws_rating.append(["State", "Program"])

    rating_data = [
        ["FL", "SS"],
        ["CA", "SS"],
        ["TX", "CW"],
        ["AL", "CW"],
        ["AZ", "SS"],
        ["GA", "SS"],
        ["NC", "CW"],
        ["TN", "SS"],
    ]

    for row in rating_data:
        ws_rating.append(row)

    # Sheet 2: AL SS Tables
    ws_ss = wb.create_sheet("AL SS Tables")
    ws_ss.append(["state", "body_type", "class", "business_class", "factor"])

    ss_data = [
        ["FL", "TRACTOR", "Class8", "AUTOHAULER", 2.15],
        ["FL", "TRACTOR", "Class8", "GENERAL", 1.00],
        ["FL", "STRAIGHT", "Class7", "GENERAL", 0.85],
        ["FL", "STRAIGHT", "Class6", "GENERAL", 0.75],
        ["CA", "TRACTOR", "Class8", "AUTOHAULER", 2.35],
        ["CA", "TRACTOR", "Class8", "GENERAL", 1.05],
        ["CA", "STRAIGHT", "Class7", "GENERAL", 0.90],
        ["AZ", "TRACTOR", "Class8", "GENERAL", 1.10],
        ["AZ", "STRAIGHT", "Class7", "GENERAL", 0.88],
        ["GA", "TRACTOR", "Class8", "AUTOHAULER", 2.10],
        ["GA", "TRACTOR", "Class8", "GENERAL", 0.98],
        ["TN", "TRACTOR", "Class8", "GENERAL", 1.02],
        ["TN", "STRAIGHT", "Class7", "GENERAL", 0.82],
    ]

    for row in ss_data:
        ws_ss.append(row)

    # Sheet 3: AL CW Tables
    ws_cw = wb.create_sheet("AL CW Tables")
    ws_cw.append(["state", "body_type", "class", "business_class", "factor"])

    cw_data = [
        ["TX", "TRACTOR", "Class8", "AUTOHAULER", 1.95],
        ["TX", "TRACTOR", "Class8", "GENERAL", 0.95],
        ["TX", "STRAIGHT", "Class7", "GENERAL", 0.80],
        ["TX", "STRAIGHT", "Class6", "GENERAL", 0.70],
        ["AL", "TRACTOR", "Class8", "AUTOHAULER", 2.05],
        ["AL", "TRACTOR", "Class8", "GENERAL", 1.00],
        ["AL", "STRAIGHT", "Class7", "GENERAL", 0.83],
        ["AL", "STRAIGHT", "Class6", "GENERAL", 0.73],
        ["NC", "TRACTOR", "Class8", "AUTOHAULER", 2.20],
        ["NC", "TRACTOR", "Class8", "GENERAL", 1.08],
        ["NC", "STRAIGHT", "Class7", "GENERAL", 0.86],
    ]

    for row in cw_data:
        ws_cw.append(row)

    # Sheet 4: Attribute Lookups
    ws_attr = wb.create_sheet("Attribute Lookups")
    ws_attr.append(["attribute_type", "category", "min_value", "max_value", "factor"])

    attr_data = [
        # Age bands
        ["age", "18-24", 18, 24, 1.50],
        ["age", "25-29", 25, 29, 1.25],
        ["age", "30-39", 30, 39, 1.00],
        ["age", "40-49", 40, 49, 0.95],
        ["age", "50-64", 50, 64, 0.90],
        ["age", "65+", 65, 100, 1.10],

        # Experience bands
        ["experience", "0-1 years", 0, 1, 1.40],
        ["experience", "2-3 years", 2, 3, 1.20],
        ["experience", "4-5 years", 4, 5, 1.05],
        ["experience", "6-10 years", 6, 10, 1.00],
        ["experience", "11+ years", 11, 50, 0.90],

        # MVR bands
        ["mvr", "0 violations", 0, 0, 0.95],
        ["mvr", "1 violation", 1, 1, 1.00],
        ["mvr", "2 violations", 2, 2, 1.15],
        ["mvr", "3 violations", 3, 3, 1.35],
        ["mvr", "4+ violations", 4, 20, 1.60],
    ]

    for row in attr_data:
        ws_attr.append(row)

    # Sheet 5: Attribute Calculations (minimal)
    ws_calc = wb.create_sheet("Attribute Calculations")
    ws_calc.append(["calculation_name", "formula", "description"])
    ws_calc.append(["driver_composite", "age_factor * exp_factor * mvr_factor", "Combined driver attributes"])
    ws_calc.append(["vehicle_composite", "body_factor * class_factor * business_factor", "Combined vehicle attributes"])

    return wb


def create_state_taxes_file():
    """Create the State Taxes and Fees 2025.xlsx file."""
    wb = Workbook()

    # Remove default sheet and create our sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    ws = wb.create_sheet("Taxes and Fees")
    ws.append(["state", "slt_pct", "stamp_pct", "fire_marshal_fee", "admitted"])

    tax_data = [
        ["FL", 0.0175, 0.0050, 5.00, True],
        ["CA", 0.0325, 0.0030, 0.00, True],
        ["TX", 0.0485, 0.0036, 75.00, False],
        ["AL", 0.0400, 0.0025, 10.00, True],
        ["AZ", 0.0350, 0.0040, 5.00, True],
        ["GA", 0.0400, 0.0025, 10.00, True],
        ["NC", 0.0425, 0.0030, 15.00, True],
        ["TN", 0.0400, 0.0020, 8.00, False],
    ]

    for row in tax_data:
        ws.append(row)

    return wb


def main():
    """Generate both Excel files."""
    data_dir = Path("/home/danielmedina/my-spec-project/romans-rater/data")

    # Ensure directory exists
    data_dir.mkdir(parents=True, exist_ok=True)

    # Create Cover Whale Rater file
    print("Creating 2025 Cover Whale Rater AL only version.xlsx...")
    cw_wb = create_cover_whale_rater_file()
    cw_file = data_dir / "2025 Cover Whale Rater AL only version.xlsx"
    cw_wb.save(cw_file)
    print(f"✓ Created: {cw_file}")
    print(f"  Sheets: {', '.join(cw_wb.sheetnames)}")

    # Create State Taxes file
    print("\nCreating State Taxes and Fees 2025.xlsx...")
    tax_wb = create_state_taxes_file()
    tax_file = data_dir / "State Taxes and Fees 2025.xlsx"
    tax_wb.save(tax_file)
    print(f"✓ Created: {tax_file}")
    print(f"  Sheets: {', '.join(tax_wb.sheetnames)}")

    print("\n✓ All sample data files created successfully!")


if __name__ == "__main__":
    main()
