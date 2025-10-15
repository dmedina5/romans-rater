"""
State tax and fee configuration loader for Roman's Rater 4.21

This module loads state-specific tax and fee rules from
"State Taxes and Fees 2025.xlsx"
"""

from pathlib import Path
from typing import Dict
import openpyxl

from ..models.factor_table import StateTaxConfiguration
from ..exceptions import LoaderError


def load_tax_config(excel_path: Path) -> Dict[str, StateTaxConfiguration]:
    """
    Load state tax and fee configuration from Excel workbook.

    Args:
        excel_path: Path to "State Taxes and Fees 2025.xlsx"

    Returns:
        Dictionary mapping state code -> StateTaxConfiguration

    Raises:
        LoaderError: If file not found, invalid format, or data validation fails
    """
    if not excel_path.exists():
        raise LoaderError(f"Tax configuration Excel file not found: {excel_path}")

    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        raise LoaderError(f"Failed to open Excel workbook: {e}")

    # Check for macro-enabled workbook (security requirement FR-047)
    if excel_path.suffix.lower() in ['.xlsm', '.xltm']:
        raise LoaderError(
            f"Macro-enabled workbook detected ({excel_path.suffix}). "
            "Only .xlsx files are allowed for security."
        )

    # Look for main sheet (could be named "State Taxes", "Taxes and Fees", etc.)
    sheet = None
    for sheet_name in workbook.sheetnames:
        if any(keyword in sheet_name.lower() for keyword in ["tax", "fee", "state"]):
            sheet = workbook[sheet_name]
            break

    if sheet is None:
        # Default to first sheet if no matching name found
        if workbook.sheetnames:
            sheet = workbook[workbook.sheetnames[0]]
        else:
            raise LoaderError("No sheets found in tax configuration workbook")

    tax_configs = _parse_tax_config_sheet(sheet)

    workbook.close()

    return tax_configs


def _parse_tax_config_sheet(sheet) -> Dict[str, StateTaxConfiguration]:
    """
    Parse tax configuration sheet.

    Expected format:
    | State | SLT % | Stamp % | Fire Marshal $ | Other Fees $ | Policy Fee Taxable | UW Fee Taxable | Broker Fee Taxable | Admitted |
    | CA    | 0.035 | 0.003   | 0.00           | 0.00         | Yes                | Yes            | Yes                | No       |
    | TX    | 0.048 | 0.006   | 1.75           | 0.00         | Yes                | Yes            | No                 | No       |

    Returns:
        Dictionary mapping state code -> StateTaxConfiguration
    """
    configs = {}

    # Find header row
    header_row = None
    col_map = {}

    for row_idx in range(1, 11):
        first_cell = sheet.cell(row_idx, 1).value
        if first_cell and "state" in str(first_cell).lower():
            header_row = row_idx
            # Build column map
            for col_idx in range(1, sheet.max_column + 1):
                header_value = sheet.cell(row_idx, col_idx).value
                if header_value:
                    header_str = str(header_value).strip().lower()
                    col_map[header_str] = col_idx
            break

    if not header_row or not col_map:
        raise LoaderError("Could not find valid header row in tax configuration sheet")

    # Validate required columns
    required_cols = ["state", "slt", "stamp"]
    for col in required_cols:
        if not any(col in key for key in col_map.keys()):
            raise LoaderError(f"Required column '{col}' not found in tax configuration sheet")

    # Parse data rows
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        state_col = _find_col_index(col_map, ["state"])
        if not state_col:
            continue

        state = sheet.cell(row_idx, state_col).value
        if not state:
            continue  # Skip empty rows

        state = str(state).strip().upper()
        if len(state) != 2:
            continue  # Skip invalid state codes

        # Parse SLT percentage
        slt_col = _find_col_index(col_map, ["slt", "surplus lines tax", "surplus"])
        slt_percentage = _parse_percentage(sheet.cell(row_idx, slt_col).value) if slt_col else 0.0

        # Parse Stamp percentage
        stamp_col = _find_col_index(col_map, ["stamp", "stamping"])
        stamp_percentage = _parse_percentage(sheet.cell(row_idx, stamp_col).value) if stamp_col else 0.0

        # Parse Fire Marshal fee
        fire_col = _find_col_index(col_map, ["fire", "marshal"])
        fire_marshal_fee = _parse_dollar_amount(sheet.cell(row_idx, fire_col).value) if fire_col else 0.0

        # Parse Other Fees
        other_col = _find_col_index(col_map, ["other", "additional"])
        other_fees = _parse_dollar_amount(sheet.cell(row_idx, other_col).value) if other_col else 0.0

        # Parse taxable fees mask
        policy_fee_taxable = _parse_boolean(
            sheet.cell(row_idx, _find_col_index(col_map, ["policy fee taxable", "policy"])).value
            if _find_col_index(col_map, ["policy fee taxable", "policy"]) else None,
            default=True
        )
        uw_fee_taxable = _parse_boolean(
            sheet.cell(row_idx, _find_col_index(col_map, ["uw fee taxable", "uw", "underwriting"])).value
            if _find_col_index(col_map, ["uw fee taxable", "uw", "underwriting"]) else None,
            default=True
        )
        broker_fee_taxable = _parse_boolean(
            sheet.cell(row_idx, _find_col_index(col_map, ["broker fee taxable", "broker"])).value
            if _find_col_index(col_map, ["broker fee taxable", "broker"]) else None,
            default=True
        )

        # Parse admitted status
        admitted_col = _find_col_index(col_map, ["admitted", "admit"])
        admitted = _parse_boolean(
            sheet.cell(row_idx, admitted_col).value if admitted_col else None,
            default=False
        )

        # Create configuration
        try:
            config = StateTaxConfiguration(
                state=state,
                slt_percentage=slt_percentage,
                stamp_percentage=stamp_percentage,
                fire_marshal_fee=fire_marshal_fee,
                other_fees=other_fees,
                taxable_fees_mask={
                    "policy_fee": policy_fee_taxable,
                    "uw_fee": uw_fee_taxable,
                    "broker_fee": broker_fee_taxable,
                },
                admitted=admitted,
            )
            configs[state] = config
        except ValueError as e:
            raise LoaderError(f"Invalid tax configuration for state {state}: {e}")

    if not configs:
        raise LoaderError("No valid tax configurations found in workbook")

    return configs


def _find_col_index(col_map: Dict[str, int], possible_names: list) -> int:
    """
    Find column index by searching for possible header names.

    Args:
        col_map: Dictionary mapping header name (lowercase) -> column index
        possible_names: List of possible column names to search for

    Returns:
        Column index (1-based) or None if not found
    """
    for name in possible_names:
        for header, col_idx in col_map.items():
            if name.lower() in header:
                return col_idx
    return None


def _parse_percentage(value) -> float:
    """
    Parse percentage value from cell.

    Handles formats like:
    - 0.035 (as decimal)
    - 3.5% (as percentage string)
    - "3.5%" (as string)

    Returns:
        Percentage as decimal (e.g., 0.035 for 3.5%)
    """
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        # If value is > 1, assume it's in percentage form (e.g., 3.5)
        if value > 1:
            return value / 100.0
        # Otherwise assume it's already decimal (e.g., 0.035)
        return float(value)

    if isinstance(value, str):
        # Remove % symbol and whitespace
        value = value.strip().replace("%", "")
        try:
            num_value = float(value)
            # If > 1, convert to decimal
            if num_value > 1:
                return num_value / 100.0
            return num_value
        except ValueError:
            return 0.0

    return 0.0


def _parse_dollar_amount(value) -> float:
    """
    Parse dollar amount from cell.

    Handles formats like:
    - 1.75 (as number)
    - "$1.75" (as string)
    - "1.75" (as string)

    Returns:
        Dollar amount as float
    """
    if value is None:
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    if isinstance(value, str):
        # Remove $ symbol, commas, and whitespace
        value = value.strip().replace("$", "").replace(",", "")
        try:
            return float(value)
        except ValueError:
            return 0.0

    return 0.0


def _parse_boolean(value, default: bool = False) -> bool:
    """
    Parse boolean value from cell.

    Handles formats like:
    - True/False (as bool)
    - "Yes"/"No" (as string)
    - "Y"/"N" (as string)
    - 1/0 (as int)

    Returns:
        Boolean value
    """
    if value is None:
        return default

    if isinstance(value, bool):
        return value

    if isinstance(value, (int, float)):
        return value != 0

    if isinstance(value, str):
        value_lower = value.strip().lower()
        if value_lower in ["yes", "y", "true", "t", "1"]:
            return True
        elif value_lower in ["no", "n", "false", "f", "0"]:
            return False

    return default
