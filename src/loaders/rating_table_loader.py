"""
Rating table loader for Roman's Rater 4.21

This module loads rating factors from the Excel workbook
"2025 Cover Whale Rater AL only version.xlsx"
"""

from datetime import date
from pathlib import Path
from typing import Dict, List, Tuple, Any
import openpyxl
from openpyxl.workbook.workbook import Workbook

from ..models.factor_table import FactorTable
from ..models.rating_program import ProgramType
from ..exceptions import LoaderError


class RatingTablesData:
    """
    Container for all rating tables loaded from Excel workbook.

    Attributes:
        rating_plan_by_state: Dict mapping state -> ProgramType (CW or SS)
        base_al_factors: Dict[(program, state)] -> base AL premium
        body_class_factors: Dict[(program, body, class, business_class)] -> factor
        radius_factors: Dict[(program, radius_bucket)] -> factor
        driver_factors: Dict[(program, age_band, exp_band, mvr_band)] -> factor
        limit_factors: Dict[(program, limit)] -> factor
        attribute_lookups: Dict with age/exp/MVR band definitions
        edition_code: Edition identifier from workbook
        rate_date: Effective date of this edition
    """

    def __init__(self):
        self.rating_plan_by_state: Dict[str, ProgramType] = {}
        self.base_al_factors: Dict[Tuple[str, str], float] = {}
        self.body_class_factors: Dict[Tuple[str, str, str, str], float] = {}
        self.radius_factors: Dict[Tuple[str, str], float] = {}
        self.driver_factors: Dict[Tuple[str, str, str, str], float] = {}
        self.limit_factors: Dict[Tuple[str, str], float] = {}
        self.attribute_lookups: Dict[str, Any] = {}
        self.edition_code: str = ""
        self.rate_date: date = date.today()


def load_rating_tables(excel_path: Path) -> RatingTablesData:
    """
    Load all rating tables from Excel workbook.

    Args:
        excel_path: Path to "2025 Cover Whale Rater AL only version.xlsx"

    Returns:
        RatingTablesData containing all parsed tables

    Raises:
        LoaderError: If file not found, required sheets missing, or data invalid
    """
    if not excel_path.exists():
        raise LoaderError(f"Rating tables Excel file not found: {excel_path}")

    try:
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        raise LoaderError(f"Failed to open Excel workbook: {e}")

    # Check for macro-enabled workbook (security requirement FR-047)
    if excel_path.suffix.lower() in ['.xlsm', '.xltm']:
        raise LoaderError(
            f"Macro-enabled workbook detected ({excel_path.suffix}). "
            "Only .xlsx files are allowed for security."
        )

    data = RatingTablesData()

    # Extract edition code and rate date from workbook metadata or first sheet
    data.edition_code = _extract_edition_code(workbook)
    data.rate_date = _extract_rate_date(workbook)

    # Load each sheet
    _load_rating_plan_by_state(workbook, data)
    _load_ss_tables(workbook, data)
    _load_cw_tables(workbook, data)
    _load_attribute_lookups(workbook, data)

    workbook.close()

    return data


def _extract_edition_code(workbook: Workbook) -> str:
    """
    Extract edition code from workbook.

    Try to find edition code in:
    1. Workbook properties
    2. First sheet cell A1 or similar
    3. Default to filename-based code
    """
    # Try to find in first sheet
    if workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]
        # Look for "Edition" or "Version" in first few rows
        for row in range(1, 6):
            for col in range(1, 4):
                cell_value = sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if "edition" in cell_value.lower() or "version" in cell_value.lower():
                        # Try to extract from next cell or same cell
                        potential_code = sheet.cell(row, col + 1).value
                        if potential_code:
                            return str(potential_code).strip()

    # Default to "2025-01" based on typical naming
    return "2025-01"


def _extract_rate_date(workbook: Workbook) -> date:
    """
    Extract rate effective date from workbook.

    Try to find date in:
    1. Workbook properties
    2. First sheet metadata cells
    3. Default to filename year
    """
    # Try to find in first sheet
    if workbook.sheetnames:
        sheet = workbook[workbook.sheetnames[0]]
        for row in range(1, 6):
            for col in range(1, 4):
                cell_value = sheet.cell(row, col).value
                if cell_value and isinstance(cell_value, str):
                    if "effective" in cell_value.lower() or "rate date" in cell_value.lower():
                        potential_date = sheet.cell(row, col + 1).value
                        if isinstance(potential_date, date):
                            return potential_date

    # Default to January 1 of year 2025 (from filename)
    return date(2025, 1, 1)


def _load_rating_plan_by_state(workbook: Workbook, data: RatingTablesData) -> None:
    """
    Load Rating Plan by State sheet.

    Expected format:
    | State | Program |
    | CA    | CW      |
    | TX    | SS      |
    """
    sheet_name = "Rating Plan by State"
    if sheet_name not in workbook.sheetnames:
        raise LoaderError(f"Required sheet '{sheet_name}' not found in workbook")

    sheet = workbook[sheet_name]

    # Find header row
    header_row = None
    for row_idx in range(1, 11):
        cell_value = sheet.cell(row_idx, 1).value
        if cell_value and "state" in str(cell_value).lower():
            header_row = row_idx
            break

    if not header_row:
        raise LoaderError(f"Could not find header row in '{sheet_name}' sheet")

    # Parse data rows
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        state = sheet.cell(row_idx, 1).value
        program = sheet.cell(row_idx, 2).value

        if not state or not program:
            continue  # Skip empty rows

        state = str(state).strip().upper()
        program = str(program).strip().upper()

        if len(state) != 2:
            raise LoaderError(f"Invalid state code: {state}")

        if program not in ["CW", "SS"]:
            raise LoaderError(f"Invalid program type: {program}. Must be CW or SS")

        data.rating_plan_by_state[state] = ProgramType(program)


def _load_ss_tables(workbook: Workbook, data: RatingTablesData) -> None:
    """
    Load AL SS Tables sheet.

    This sheet contains factor tables for SS (Specialty Standard) program.
    Expected structure varies by section (Base AL, Body/Class, Radius, Driver, Limit).
    """
    sheet_name = "AL SS Tables"
    if sheet_name not in workbook.sheetnames:
        raise LoaderError(f"Required sheet '{sheet_name}' not found in workbook")

    sheet = workbook[sheet_name]

    # Parse different sections of the sheet
    # Each section has a header identifying the factor type
    _parse_factor_sections(sheet, "SS", data)


def _load_cw_tables(workbook: Workbook, data: RatingTablesData) -> None:
    """
    Load AL CW Tables sheet.

    This sheet contains factor tables for CW (Cover Whale) program.
    Expected structure matches SS tables format.
    """
    sheet_name = "AL CW Tables"
    if sheet_name not in workbook.sheetnames:
        raise LoaderError(f"Required sheet '{sheet_name}' not found in workbook")

    sheet = workbook[sheet_name]

    # Parse different sections of the sheet
    _parse_factor_sections(sheet, "CW", data)


def _parse_factor_sections(sheet, program: str, data: RatingTablesData) -> None:
    """
    Parse factor sections from a tables sheet.

    Looks for section headers like:
    - "Base AL" or "Base AL Premium"
    - "Body/Class Factors"
    - "Radius Factors"
    - "Driver Factors"
    - "Limit Factors"
    """
    current_section = None

    for row_idx in range(1, sheet.max_row + 1):
        # Check for section headers
        first_cell = sheet.cell(row_idx, 1).value
        if not first_cell:
            continue

        first_cell_str = str(first_cell).strip().lower()

        # Identify section
        if "base al" in first_cell_str:
            current_section = "base_al"
            continue
        elif "body" in first_cell_str and "class" in first_cell_str:
            current_section = "body_class"
            continue
        elif "radius" in first_cell_str:
            current_section = "radius"
            continue
        elif "driver" in first_cell_str:
            current_section = "driver"
            continue
        elif "limit" in first_cell_str:
            current_section = "limit"
            continue

        # Parse data based on current section
        if current_section == "base_al":
            _parse_base_al_row(sheet, row_idx, program, data)
        elif current_section == "body_class":
            _parse_body_class_row(sheet, row_idx, program, data)
        elif current_section == "radius":
            _parse_radius_row(sheet, row_idx, program, data)
        elif current_section == "driver":
            _parse_driver_row(sheet, row_idx, program, data)
        elif current_section == "limit":
            _parse_limit_row(sheet, row_idx, program, data)


def _parse_base_al_row(sheet, row_idx: int, program: str, data: RatingTablesData) -> None:
    """Parse Base AL row: State | Base AL Premium"""
    state = sheet.cell(row_idx, 1).value
    base_al = sheet.cell(row_idx, 2).value

    if not state or not isinstance(base_al, (int, float)):
        return  # Skip invalid rows

    state = str(state).strip().upper()
    if len(state) == 2:  # Valid state code
        data.base_al_factors[(program, state)] = float(base_al)


def _parse_body_class_row(sheet, row_idx: int, program: str, data: RatingTablesData) -> None:
    """Parse Body/Class row: Body | Class | Business Class | Factor"""
    body = sheet.cell(row_idx, 1).value
    vehicle_class = sheet.cell(row_idx, 2).value
    business_class = sheet.cell(row_idx, 3).value
    factor = sheet.cell(row_idx, 4).value

    if not all([body, vehicle_class, business_class]) or not isinstance(factor, (int, float)):
        return  # Skip invalid rows

    key = (
        program,
        str(body).strip(),
        str(vehicle_class).strip(),
        str(business_class).strip()
    )
    data.body_class_factors[key] = float(factor)


def _parse_radius_row(sheet, row_idx: int, program: str, data: RatingTablesData) -> None:
    """Parse Radius row: Radius Bucket | Factor"""
    radius_bucket = sheet.cell(row_idx, 1).value
    factor = sheet.cell(row_idx, 2).value

    if not radius_bucket or not isinstance(factor, (int, float)):
        return  # Skip invalid rows

    key = (program, str(radius_bucket).strip())
    data.radius_factors[key] = float(factor)


def _parse_driver_row(sheet, row_idx: int, program: str, data: RatingTablesData) -> None:
    """Parse Driver row: Age Band | Experience Band | MVR Band | Factor"""
    age_band = sheet.cell(row_idx, 1).value
    exp_band = sheet.cell(row_idx, 2).value
    mvr_band = sheet.cell(row_idx, 3).value
    factor = sheet.cell(row_idx, 4).value

    if not all([age_band, exp_band, mvr_band]) or not isinstance(factor, (int, float)):
        return  # Skip invalid rows

    key = (
        program,
        str(age_band).strip(),
        str(exp_band).strip(),
        str(mvr_band).strip()
    )
    data.driver_factors[key] = float(factor)


def _parse_limit_row(sheet, row_idx: int, program: str, data: RatingTablesData) -> None:
    """Parse Limit row: Limit | Factor"""
    limit = sheet.cell(row_idx, 1).value
    factor = sheet.cell(row_idx, 2).value

    if not limit or not isinstance(factor, (int, float)):
        return  # Skip invalid rows

    key = (program, str(limit).strip())
    data.limit_factors[key] = float(factor)


def _load_attribute_lookups(workbook: Workbook, data: RatingTablesData) -> None:
    """
    Load Attribute Lookups sheet.

    This sheet defines band boundaries for:
    - Age bands (e.g., "18-24", "25-34", etc.)
    - Experience bands (e.g., "0-2", "3-5", etc.)
    - MVR bands (e.g., "0", "1-2", "3+", etc.)
    """
    sheet_name = "Attribute Lookups"
    if sheet_name not in workbook.sheetnames:
        # This sheet is optional; use default bands if missing
        data.attribute_lookups = _get_default_attribute_bands()
        return

    sheet = workbook[sheet_name]

    # Parse age bands
    age_bands = _parse_attribute_band_section(sheet, "age")
    if age_bands:
        data.attribute_lookups["age_bands"] = age_bands

    # Parse experience bands
    exp_bands = _parse_attribute_band_section(sheet, "experience")
    if exp_bands:
        data.attribute_lookups["experience_bands"] = exp_bands

    # Parse MVR bands
    mvr_bands = _parse_attribute_band_section(sheet, "mvr")
    if mvr_bands:
        data.attribute_lookups["mvr_bands"] = mvr_bands

    # Use defaults for any missing bands
    defaults = _get_default_attribute_bands()
    for key, value in defaults.items():
        if key not in data.attribute_lookups:
            data.attribute_lookups[key] = value


def _parse_attribute_band_section(sheet, attribute_type: str) -> List[Dict[str, Any]]:
    """
    Parse an attribute band section from the sheet.

    Returns list of band definitions like:
    [
        {"label": "18-24", "min": 18, "max": 24},
        {"label": "25-34", "min": 25, "max": 34},
        ...
    ]
    """
    bands = []

    # Find section header
    header_row = None
    for row_idx in range(1, sheet.max_row + 1):
        cell_value = sheet.cell(row_idx, 1).value
        if cell_value and attribute_type in str(cell_value).lower():
            header_row = row_idx
            break

    if not header_row:
        return bands

    # Parse band rows (Label | Min | Max)
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        label = sheet.cell(row_idx, 1).value
        min_val = sheet.cell(row_idx, 2).value
        max_val = sheet.cell(row_idx, 3).value

        if not label:
            break  # End of section

        band = {
            "label": str(label).strip(),
            "min": min_val if isinstance(min_val, (int, float)) else None,
            "max": max_val if isinstance(max_val, (int, float)) else None,
        }
        bands.append(band)

    return bands


def _get_default_attribute_bands() -> Dict[str, List[Dict[str, Any]]]:
    """
    Return default attribute band definitions.

    These are used as fallback if Attribute Lookups sheet is missing or incomplete.
    """
    return {
        "age_bands": [
            {"label": "18-24", "min": 18, "max": 24},
            {"label": "25-34", "min": 25, "max": 34},
            {"label": "35-49", "min": 35, "max": 49},
            {"label": "50-64", "min": 50, "max": 64},
            {"label": "65+", "min": 65, "max": None},
        ],
        "experience_bands": [
            {"label": "0-2", "min": 0, "max": 2},
            {"label": "3-5", "min": 3, "max": 5},
            {"label": "6-10", "min": 6, "max": 10},
            {"label": "11+", "min": 11, "max": None},
        ],
        "mvr_bands": [
            {"label": "0", "min": 0, "max": 0},
            {"label": "1-2", "min": 1, "max": 2},
            {"label": "3-4", "min": 3, "max": 4},
            {"label": "5+", "min": 5, "max": None},
        ],
    }
