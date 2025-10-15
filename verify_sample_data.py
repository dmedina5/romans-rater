#!/usr/bin/env python3
"""
Verify the structure of the generated Excel files.
"""

from openpyxl import load_workbook
from pathlib import Path


def verify_cover_whale_rater():
    """Verify the Cover Whale Rater file structure."""
    file_path = Path("/home/danielmedina/my-spec-project/romans-rater/data/2025 Cover Whale Rater AL only version.xlsx")

    print("=" * 70)
    print("Verifying: 2025 Cover Whale Rater AL only version.xlsx")
    print("=" * 70)

    wb = load_workbook(file_path)
    print(f"\nSheets: {wb.sheetnames}\n")

    # Check each sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{sheet_name}:")
        print(f"  Rows: {ws.max_row}")
        print(f"  Columns: {ws.max_column}")

        # Print headers
        headers = [cell.value for cell in ws[1]]
        print(f"  Headers: {headers}")

        # Print first 3 data rows
        if ws.max_row > 1:
            print(f"  Sample data (first 3 rows):")
            for i in range(2, min(5, ws.max_row + 1)):
                row_data = [cell.value for cell in ws[i]]
                print(f"    Row {i}: {row_data}")


def verify_state_taxes():
    """Verify the State Taxes file structure."""
    file_path = Path("/home/danielmedina/my-spec-project/romans-rater/data/State Taxes and Fees 2025.xlsx")

    print("\n" + "=" * 70)
    print("Verifying: State Taxes and Fees 2025.xlsx")
    print("=" * 70)

    wb = load_workbook(file_path)
    print(f"\nSheets: {wb.sheetnames}\n")

    # Check the sheet
    ws = wb["Taxes and Fees"]
    print(f"Taxes and Fees:")
    print(f"  Rows: {ws.max_row}")
    print(f"  Columns: {ws.max_column}")

    # Print headers
    headers = [cell.value for cell in ws[1]]
    print(f"  Headers: {headers}")

    # Print all data rows
    if ws.max_row > 1:
        print(f"  Data (all rows):")
        for i in range(2, ws.max_row + 1):
            row_data = [cell.value for cell in ws[i]]
            print(f"    Row {i}: {row_data}")


def main():
    """Verify both files."""
    verify_cover_whale_rater()
    verify_state_taxes()
    print("\n" + "=" * 70)
    print("Verification complete!")
    print("=" * 70)


if __name__ == "__main__":
    main()
